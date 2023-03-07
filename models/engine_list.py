from string import ascii_lowercase
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from models.engine import Engine


class EngineList:
    def __init__(self, source_filepath: str):
        self._init_data_from_file(source_filepath)
        self._init_shift_freqs_list()
        self._init_power_list()
        self._init_engine_list()

	#? [INITIALIZATION] Get excel book and sheet from file
    def _init_data_from_file(self, filename: str):
        self._engine_book: Workbook = load_workbook(filename)
        self._sheet: Worksheet = self._engine_book.active

	#? [INITIALIZATION] Get constant shift rotation freqs from excel table
    def _init_shift_freqs_list(self):
        self._constant_shift_freqs = tuple(sorted(set(
            col[0] for col in self._sheet.iter_cols(values_only=True) if isinstance(col[0], float)
        )))

	#? [INITIALIZATION] Get constant power list from excel table
    def _init_power_list(self):
        self._constant_power_list = tuple(sorted(
            row[0] for row in self._sheet.iter_rows(values_only=True) if isinstance(row[0], float)
        ))

	#? [INITIALIZATION] Get engine list from excel table
    def _init_engine_list(self):
        engine_columns = tuple(
            col_num for col_num, col in enumerate(self._sheet.iter_cols(values_only=True)) if col[1] == 'Тип двигателя'
        )
        self._engine_list: list[Engine] = []
        for eng_col in engine_columns:
            for row_num in range(2, self._sheet.max_row):
                engine_title = self._sheet[ascii_lowercase[eng_col] + str(row_num + 1)].value
                if engine_title is not None:
                    self._engine_list.append(Engine(
                        title=engine_title,
                        power=self._sheet['a' + str(row_num + 1)].value,
                        rotation_freq=self._sheet[ascii_lowercase[eng_col + 1] + str(row_num + 1)].value,
                        shift_freq=self._sheet[ascii_lowercase[eng_col] + '1'].value,
                        delta_t=self._sheet[ascii_lowercase[eng_col + 2] + str(row_num + 1)].value,
                    ))

	#* ------------- GETTERS -------------
    def get_constant_shift_freqs(self):
        return self._constant_shift_freqs

    def get_constant_power_list(self):
        return self._constant_power_list

    def get_engine_list(self):
        return self._engine_list[:]
	#* -----------------------------------

	#? Try to find the right engine by current [shift rotation freq] and current [power]
    def find_required_engine(self, current_freq: float, current_power: float):
        closest_constant_freq = min(
            shift_freq for shift_freq in self._constant_shift_freqs if (shift_freq - current_freq >= 0)
        )

        closest_constant_power = min(
            power for power in self._constant_power_list if (power - current_power >= 0)
        )
        return tuple(filter(
            lambda engine: engine.shift_freq == closest_constant_freq and engine.power == closest_constant_power,
            self._engine_list,
        ))[0]
